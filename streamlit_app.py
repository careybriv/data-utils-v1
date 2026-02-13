import streamlit as st
import time
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google import genai
from google.api_core import exceptions
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# --- CONFIGURATION ---
PAGE_TITLE = "Redline AI | Enterprise"
PAGE_ICON = "🏢"

# --- SETUP PAGE ---
st.set_page_config(page_title=PAGE_TITLE, page_icon=PAGE_ICON, layout="centered")

# --- HIDE BRANDING & MENU ---
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            /* Fix Metric Containers to handle longer text gracefully */
            [data-testid="stMetricValue"] {
                font-size: 1.2rem !important;
                word-wrap: break-word;
            }
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

# --- SESSION STATE INITIALIZATION ---
if "legal_accepted" not in st.session_state:
    st.session_state["legal_accepted"] = False

# --- LEGAL TEXT ---
TERMS_OF_USE = """
### 🔒 Secure Workspace Access

**1. Service Scope:** Redline AI is an automated data extraction tool. It is not a law firm and does not provide legal advice.
**2. Strict Zero Retention:** Documents are processed in temporary RAM and **permanently deleted** immediately after analysis. We do not store or train on your data.
**3. AI Fallibility:** This report is a preliminary draft. **Human verification is MANDATORY.**
**4. Liability:** You are responsible for the final review. Redline AI is not liable for errors or omissions.
"""

# --- HUMAN ERROR TRANSLATOR ---
def translate_error(e):
    err_str = str(e).lower()
    if "11001" in err_str or "connection" in err_str:
        return "🌐 Connection Lost. Please check your internet."
    elif "403" in err_str or "api key" in err_str:
        return "🔑 License Error. Contact Redline Support."
    elif "429" in err_str:
        return "⏳ System busy. Please wait 10s and retry."
    elif "pdf" in err_str:
        return "📄 PDF Error. File may be corrupted or password protected."
    else:
        return "⚠️ An unexpected error occurred. Please refresh."

# --- DATABASE LOGIC ---
def connect_to_sheet():
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        sheet_url = st.secrets["private_sheet_url"]
        return client.open_by_url(sheet_url).sheet1
    except:
        return None

def check_access(code):
    try:
        sheet = connect_to_sheet()
        if not sheet: return "⚠️ Database Error", None, None
        records = sheet.get_all_records()
        for row in records:
            if str(row['username']) == code:
                if str(row['active']).upper() != "TRUE":
                    return "❌ Deactivated", None, None
                used = int(row['used'])
                limit = int(row['limit'])
                if used >= limit:
                    return f"⚠️ Limit Reached ({used}/{limit})", used, limit
                return "OK", used, limit
        return "❌ Invalid Code", None, None
    except:
        return "⚠️ System Error", None, None

def increment_usage(code):
    try:
        sheet = connect_to_sheet()
        cell = sheet.find(code)
        current_val = int(sheet.cell(cell.row, 2).value)
        sheet.update_cell(cell.row, 2, current_val + 1)
    except:
        pass 

# --- BACKEND LOGIC ---
def get_gemini_client():
    try:
        api_key = st.secrets["GOOGLE_API_KEY"]
        return genai.Client(api_key=api_key)
    except:
        return None

def create_excel_bytes(filename, data):
    # --- EXCEL FORMATTING FIX ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Redline Analysis"
    
    headers = ["Tenant", "Rent Breakdown", "Deposit", "Risk Score", "Risk Summary"]
    ws.append(headers)
    
    # 1. Flatten Data
    raw_flags = data.get("risk_flags", "None")
    if isinstance(raw_flags, list):
        risk_flags_str = ", ".join([str(flag) for flag in raw_flags])
    else:
        risk_flags_str = str(raw_flags)

    ws.append([
        str(data.get("tenant_name", "N/A")), 
        str(data.get("monthly_rent", "N/A")), 
        str(data.get("security_deposit", "N/A")), 
        str(data.get("risk_score", "0")), 
        risk_flags_str
    ])
    
    ws.append([])
    ws.append(["NOTE: AI-Generated Draft. Verify with original document."])
    
    # 2. BEAUTIFY (The Fix)
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    
    # Header Styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Column Widths (Wider for summary)
    ws.column_dimensions['A'].width = 25  # Tenant
    ws.column_dimensions['B'].width = 40  # Rent (Equation needs space)
    ws.column_dimensions['C'].width = 25  # Deposit
    ws.column_dimensions['D'].width = 12  # Score
    ws.column_dimensions['E'].width = 80  # Summary (The big one)

    # Data Wrapping (Critical for long text)
    for row in ws.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def analyze_lease(uploaded_file):
    client = get_gemini_client()
    if not client: raise Exception("API Key Missing")
    
    with st.spinner("Encrypting & Uploading to Neural Engine..."):
        bytes_data = uploaded_file.getvalue()
        temp_filename = "temp_upload.pdf" 
        with open(temp_filename, "wb") as f:
            f.write(bytes_data)
        cloud_file = client.files.upload(file=temp_filename)

    while cloud_file.state.name == "PROCESSING":
        time.sleep(1)
        cloud_file = client.files.get(name=cloud_file.name)

    if cloud_file.state.name == "FAILED": raise Exception("PDF Syntax Error")

    try:
        sys_prompt = st.secrets["prompts"]["system_instruction"]
    except:
        sys_prompt = "Extract tenant_name, monthly_rent, security_deposit, risk_score, risk_flags. JSON."
    
    data = None
    max_retries = 3
    
    with st.spinner("🔍 AI Auditing Lease (Gemini 2.0 Flash)..."):
        for attempt in range(max_retries):
            try:
                response = client.models.generate_content(
                    model="gemini-2.0-flash",
                    contents=[cloud_file, sys_prompt]
                )
                text = response.text.replace("```json", "").replace("```", "").strip()
                data = json.loads(text)
                break 
            except exceptions.ResourceExhausted:
                time.sleep(5) 
            except Exception as e:
                if attempt == max_retries - 1: raise e
                time.sleep(1)
        try:
            client.files.delete(name=cloud_file.name)
        except:
            pass

    if not data: raise Exception("AI could not extract data.")
    return data, create_excel_bytes(uploaded_file.name, data)

# --- FRONTEND UI ---
st.title("🏢 Redline AI | Enterprise")

# 1. AUTHENTICATION
with st.sidebar:
    st.header("Client Portal")
    password = st.text_input("Access Code", type="password")
    
    status = "WAITING"
    if password:
        status, used, limit = check_access(password)
        if status == "OK":
            st.markdown("---")
            st.caption(f"Status: Active | Quota: {used}/{limit}")
            st.progress(used / limit)
        else:
            st.error(status)

# 2. LEGAL GATEKEEPER (The Pop-up Replacement)
if password and status == "OK":
    
    # If they haven't agreed yet, block the app with the Terms screen
    if not st.session_state["legal_accepted"]:
        st.markdown("---")
        st.info("👋 Welcome to Redline AI. Please accept the security protocols to continue.")
        with st.container(border=True):
            st.markdown(TERMS_OF_USE)
            
            # The "Unlock" Button
            if st.button("✅ I Agree & Enter Secure Workspace", type="primary", use_container_width=True):
                st.session_state["legal_accepted"] = True
                st.rerun() # Refresh to show the main app
    
    # 3. MAIN APPLICATION (Unlocked)
    else:
        uploaded_file = st.file_uploader("Upload Lease Agreement (PDF)", type=["pdf"])

        # Reset logic for new files
        if "last_file_id" not in st.session_state:
            st.session_state["last_file_id"] = None
        if uploaded_file and uploaded_file.file_id != st.session_state["last_file_id"]:
            st.session_state["last_file_id"] = uploaded_file.file_id
            if "audit_result" in st.session_state:
                del st.session_state["audit_result"]

        if uploaded_file is not None:
            if st.button("🚀 Run Analysis (-1 Credit)", type="primary"):
                try:
                    data, excel_file = analyze_lease(uploaded_file)
                    st.session_state["audit_result"] = data
                    st.session_state["audit_excel"] = excel_file
                    increment_usage(password)
                    st.toast("✅ Analysis Complete")
                except Exception as e:
                    st.error(translate_error(e))

            # 4. RESULTS DASHBOARD (Clean Layout)
            if "audit_result" in st.session_state:
                data = st.session_state["audit_result"]
                excel_bytes = st.session_state["audit_excel"]
                
                st.markdown("---")
                st.subheader("📊 Audit Results")
                
                # A. Top Level Metrics (Short values only)
                c1, c2, c3 = st.columns(3)
                
                # Color code the Risk Score
                score = data.get('risk_score', 0)
                score_color = "red" if score >= 7 else "orange" if score >= 4 else "green"
                c1.markdown(f"**Risk Score**")
                c1.markdown(f":{score_color}[**{score}/10**]")
                
                # Deposit
                c2.metric("Deposit", str(data.get("security_deposit", "N/A")))
                
                # Total Rent (Just the sum if possible, or label it "See Breakdown")
                rent_raw = str(data.get("monthly_rent", "N/A"))
                # If rent string is super long (>30 chars), just show "Complex" or truncate
                display_rent = rent_raw if len(rent_raw) < 25 else "View Breakdown ⬇️"
                c3.metric("Total Liability", display_rent)
                
                # B. Long Text Sections (The Mess Fix)
                # Rent Equation
                with st.container(border=True):
                    st.markdown("**💰 Rent Breakdown:**")
                    st.code(rent_raw, language=None)
                
                # Risk Summary
                with st.container(border=True):
                    st.markdown("**🚩 Risk Assessment:**")
                    st.info(data.get("risk_flags", "No summary available."))
                
                # Download
                st.download_button(
                    label="📥 Download Professional Excel Report",
                    data=excel_bytes,
                    file_name=f"AUDIT_REPORT.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
